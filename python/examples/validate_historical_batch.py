#!/usr/bin/env python3
"""Historical batch validation against market sample and QuantLib reference."""

from __future__ import annotations

import argparse
import ctypes as ct
import json
import math
import os
import subprocess
import sys
import time
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parents[2]
PYTHON_DIR = ROOT / "python"
if str(PYTHON_DIR) not in sys.path:
    sys.path.insert(0, str(PYTHON_DIR))

os.environ.setdefault("QK_LIB_PATH", str(ROOT / "build" / "cpp"))

from quantkernel import QK_CALL, QK_PUT, QuantKernel  # noqa: E402


@dataclass
class OptionBatch:
    spot: np.ndarray
    strike: np.ndarray
    tau: np.ndarray
    vol: np.ndarray
    bid: np.ndarray
    ask: np.ndarray
    option_type: np.ndarray
    r: np.ndarray
    q: np.ndarray

    @property
    def size(self) -> int:
        return int(self.spot.shape[0])


def _float_or_none(value) -> float | None:
    try:
        x = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(x):
        return None
    return x


def _infer_q_from_parity(
    spot: float,
    strike: float,
    tau: float,
    call_mid: float,
    put_mid: float,
    risk_free_rate: float,
) -> float | None:
    """Infer dividend yield q from put-call parity with fixed r.

    Parity: C - P = S*exp(-qT) - K*exp(-rT)
    """
    lhs = call_mid - put_mid + strike * math.exp(-risk_free_rate * tau)
    if lhs <= 0.0:
        return None
    ratio = lhs / spot
    if ratio <= 0.0:
        return None
    q = -math.log(ratio) / tau
    if not math.isfinite(q):
        return None
    return q


def _collect_batches(
    xlsx_path: Path,
    sheet_name: str | None,
    dte_basis: float,
    carry_mode: str,
    risk_free_rate: float,
    dividend_yield: float,
    pcp_q_abs_max: float,
    sample_per_side: int,
    seed: int,
) -> tuple[OptionBatch, OptionBatch, dict[str, int]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    idx = {name: i for i, name in enumerate(header)}
    required = [
        "UNDERLYING_LAST",
        "DTE",
        "STRIKE",
        "C_IV",
        "C_BID",
        "C_ASK",
        "P_IV",
        "P_BID",
        "P_ASK",
    ]
    missing = [name for name in required if name not in idx]
    if missing:
        raise ValueError(f"Missing required columns in sheet '{ws.title}': {missing}")

    call_rows: list[tuple[float, float, float, float, float, float, float, float]] = []
    put_rows: list[tuple[float, float, float, float, float, float, float, float]] = []
    stats = {
        "rows_seen": 0,
        "rows_with_valid_core_fields": 0,
        "valid_call_rows": 0,
        "valid_put_rows": 0,
        "valid_call_put_rows": 0,
        "pcp_usable_rows": 0,
        "pcp_rejected_nonpositive_lhs": 0,
        "pcp_rejected_abs_q_cap": 0,
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats["rows_seen"] += 1
        s = _float_or_none(row[idx["UNDERLYING_LAST"]])
        dte = _float_or_none(row[idx["DTE"]])
        k = _float_or_none(row[idx["STRIKE"]])
        if s is None or dte is None or k is None or s <= 0.0 or dte <= 0.0 or k <= 0.0:
            continue
        t = dte / dte_basis
        if not (math.isfinite(t) and t > 0.0):
            continue

        stats["rows_with_valid_core_fields"] += 1

        c_iv = _float_or_none(row[idx["C_IV"]])
        c_bid = _float_or_none(row[idx["C_BID"]])
        c_ask = _float_or_none(row[idx["C_ASK"]])
        call_valid = (
            c_iv is not None
            and c_iv > 0.0
            and c_bid is not None
            and c_ask is not None
            and c_bid >= 0.0
            and c_ask >= 0.0
            and c_ask >= c_bid
        )
        if call_valid:
            stats["valid_call_rows"] += 1

        p_iv = _float_or_none(row[idx["P_IV"]])
        p_bid = _float_or_none(row[idx["P_BID"]])
        p_ask = _float_or_none(row[idx["P_ASK"]])
        put_valid = (
            p_iv is not None
            and p_iv > 0.0
            and p_bid is not None
            and p_ask is not None
            and p_bid >= 0.0
            and p_ask >= 0.0
            and p_ask >= p_bid
        )
        if put_valid:
            stats["valid_put_rows"] += 1

        if call_valid and put_valid:
            stats["valid_call_put_rows"] += 1

        if carry_mode == "flat":
            if call_valid:
                call_rows.append((s, k, t, c_iv, c_bid, c_ask, risk_free_rate, dividend_yield))
            if put_valid:
                put_rows.append((s, k, t, p_iv, p_bid, p_ask, risk_free_rate, dividend_yield))
            continue

        if not (call_valid and put_valid):
            continue

        call_mid = 0.5 * (c_bid + c_ask)
        put_mid = 0.5 * (p_bid + p_ask)
        q_pcp = _infer_q_from_parity(s, k, t, call_mid, put_mid, risk_free_rate)
        if q_pcp is None:
            stats["pcp_rejected_nonpositive_lhs"] += 1
            continue
        if abs(q_pcp) > pcp_q_abs_max:
            stats["pcp_rejected_abs_q_cap"] += 1
            continue

        stats["pcp_usable_rows"] += 1
        call_rows.append((s, k, t, c_iv, c_bid, c_ask, risk_free_rate, q_pcp))
        put_rows.append((s, k, t, p_iv, p_bid, p_ask, risk_free_rate, q_pcp))

    rng = np.random.default_rng(seed)

    def _make_batch(
        rows: list[tuple[float, float, float, float, float, float, float, float]], opt_type: int
    ) -> OptionBatch:
        if sample_per_side > 0 and len(rows) > sample_per_side:
            picks = rng.choice(len(rows), size=sample_per_side, replace=False)
            rows = [rows[i] for i in np.sort(picks)]
        arr = np.asarray(rows, dtype=np.float64)
        if arr.size == 0:
            return OptionBatch(
                np.empty(0, dtype=np.float64),
                np.empty(0, dtype=np.float64),
                np.empty(0, dtype=np.float64),
                np.empty(0, dtype=np.float64),
                np.empty(0, dtype=np.float64),
                np.empty(0, dtype=np.float64),
                np.empty(0, dtype=np.int32),
                np.empty(0, dtype=np.float64),
                np.empty(0, dtype=np.float64),
            )
        return OptionBatch(
            spot=np.ascontiguousarray(arr[:, 0]),
            strike=np.ascontiguousarray(arr[:, 1]),
            tau=np.ascontiguousarray(arr[:, 2]),
            vol=np.ascontiguousarray(arr[:, 3]),
            bid=np.ascontiguousarray(arr[:, 4]),
            ask=np.ascontiguousarray(arr[:, 5]),
            option_type=np.full(arr.shape[0], opt_type, dtype=np.int32),
            r=np.ascontiguousarray(arr[:, 6]),
            q=np.ascontiguousarray(arr[:, 7]),
        )

    return _make_batch(call_rows, QK_CALL), _make_batch(put_rows, QK_PUT), stats


def _compile_quantlib_wrapper(
    wrapper_src: Path,
    wrapper_lib: Path,
    quantlib_source_root: Path,
    quantlib_build_dir: Path,
) -> None:
    wrapper_lib.parent.mkdir(parents=True, exist_ok=True)
    quantlib_lib_dir = quantlib_build_dir / "ql"
    quantlib_lib = quantlib_lib_dir / "libQuantLib.so"
    if not quantlib_lib.exists():
        raise FileNotFoundError(f"QuantLib shared library not found at {quantlib_lib}")
    if not wrapper_src.exists():
        raise FileNotFoundError(f"QuantLib wrapper source not found at {wrapper_src}")

    needs_build = (
        not wrapper_lib.exists()
        or wrapper_lib.stat().st_mtime < wrapper_src.stat().st_mtime
        or wrapper_lib.stat().st_mtime < quantlib_lib.stat().st_mtime
    )
    if not needs_build:
        return

    cmd = [
        "g++",
        "-O3",
        "-std=c++17",
        "-fPIC",
        "-shared",
        str(wrapper_src),
        "-I",
        str(quantlib_source_root),
        "-L",
        str(quantlib_lib_dir),
        "-lQuantLib",
        f"-Wl,-rpath,{quantlib_lib_dir}",
        "-o",
        str(wrapper_lib),
    ]
    subprocess.run(cmd, check=True)


class QuantLibBSMRef:
    def __init__(self, lib_path: Path):
        self._lib = ct.CDLL(str(lib_path))
        self._fn = self._lib.ql_ref_black_scholes_merton_price_batch
        dptr = ct.POINTER(ct.c_double)
        i32ptr = ct.POINTER(ct.c_int32)
        self._fn.argtypes = [
            dptr,
            dptr,
            dptr,
            dptr,
            dptr,
            dptr,
            i32ptr,
            ct.c_int32,
            dptr,
            ct.POINTER(ct.c_char),
            ct.c_int32,
        ]
        self._fn.restype = ct.c_int32

    def price_batch(
        self,
        spot: np.ndarray,
        strike: np.ndarray,
        tau: np.ndarray,
        vol: np.ndarray,
        r: np.ndarray,
        q: np.ndarray,
        option_type: np.ndarray,
    ) -> np.ndarray:
        n = int(spot.shape[0])
        out = np.empty(n, dtype=np.float64)
        err = ct.create_string_buffer(1024)
        rc = self._fn(
            spot.ctypes.data_as(ct.POINTER(ct.c_double)),
            strike.ctypes.data_as(ct.POINTER(ct.c_double)),
            tau.ctypes.data_as(ct.POINTER(ct.c_double)),
            vol.ctypes.data_as(ct.POINTER(ct.c_double)),
            r.ctypes.data_as(ct.POINTER(ct.c_double)),
            q.ctypes.data_as(ct.POINTER(ct.c_double)),
            option_type.ctypes.data_as(ct.POINTER(ct.c_int32)),
            n,
            out.ctypes.data_as(ct.POINTER(ct.c_double)),
            err,
            len(err),
        )
        if rc != 0:
            msg = err.value.decode("utf-8", errors="replace").strip() or f"rc={rc}"
            raise RuntimeError(f"QuantLib wrapper failed: {msg}")
        return out


def _market_metrics(model: np.ndarray, bid: np.ndarray, ask: np.ndarray) -> dict[str, float]:
    mid = 0.5 * (bid + ask)
    err = model - mid
    abs_err = np.abs(err)
    spread = ask - bid

    below = np.maximum(bid - model, 0.0)
    above = np.maximum(model - ask, 0.0)
    outside = below + above
    inside = outside == 0.0

    return {
        "n": float(model.size),
        "mae_to_mid": float(np.mean(abs_err)),
        "rmse_to_mid": float(np.sqrt(np.mean(err * err))),
        "p95_abs_error_to_mid": float(np.percentile(abs_err, 95)),
        "p99_abs_error_to_mid": float(np.percentile(abs_err, 99)),
        "inside_bid_ask_ratio": float(np.mean(inside)),
        "mean_outside_bid_ask": float(np.mean(outside)),
        "p95_outside_bid_ask": float(np.percentile(outside, 95)),
        "mean_bid_ask_spread": float(np.mean(spread)),
    }


def _pair_metrics(a: np.ndarray, b: np.ndarray) -> dict[str, float]:
    diff = a - b
    abs_diff = np.abs(diff)
    return {
        "n": float(a.size),
        "mae": float(np.mean(abs_diff)),
        "rmse": float(np.sqrt(np.mean(diff * diff))),
        "p95_abs": float(np.percentile(abs_diff, 95)),
        "p99_abs": float(np.percentile(abs_diff, 99)),
        "max_abs": float(np.max(abs_diff)),
    }


def _run_side(
    qk: QuantKernel,
    ql_ref: QuantLibBSMRef,
    side_name: str,
    batch: OptionBatch,
) -> dict:
    if batch.size == 0:
        return {"side": side_name, "n": 0}

    t0 = time.perf_counter()
    qk_price = qk.black_scholes_merton_price_batch(
        batch.spot, batch.strike, batch.tau, batch.vol, batch.r, batch.q, batch.option_type
    )
    t1 = time.perf_counter()

    ql_price = ql_ref.price_batch(
        batch.spot, batch.strike, batch.tau, batch.vol, batch.r, batch.q, batch.option_type
    )
    t2 = time.perf_counter()

    return {
        "side": side_name,
        "n": batch.size,
        "timing_seconds": {
            "quantkernel_batch": float(t1 - t0),
            "quantlib_reference_batch": float(t2 - t1),
        },
        "quantkernel_vs_market": _market_metrics(qk_price, batch.bid, batch.ask),
        "quantlib_vs_market": _market_metrics(ql_price, batch.bid, batch.ask),
        "quantkernel_vs_quantlib": _pair_metrics(qk_price, ql_price),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=ROOT / "third_party_source" / "option_price_sample.xlsx",
        help="Path to option sample workbook.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Sheet name. Defaults to first sheet.",
    )
    parser.add_argument(
        "--dte-basis",
        type=float,
        default=365.0,
        help="Convert DTE to year fraction by dividing by this value.",
    )
    parser.add_argument(
        "--risk-free-rate",
        type=float,
        default=0.05,
        help="Flat risk-free rate used in flat mode and as PCP discount anchor.",
    )
    parser.add_argument(
        "--dividend-yield",
        type=float,
        default=0.015,
        help="Flat dividend yield used in flat carry mode.",
    )
    parser.add_argument(
        "--carry-mode",
        type=str,
        choices=["flat", "pcp"],
        default="flat",
        help="Carry assumption: flat r/q or PCP-inferred per-row q (with fixed r).",
    )
    parser.add_argument(
        "--pcp-q-abs-max",
        type=float,
        default=5.0,
        help="Reject PCP-inferred rows with |q| above this threshold.",
    )
    parser.add_argument(
        "--sample-per-side",
        type=int,
        default=0,
        help="If > 0, random subsample size per call/put side.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="RNG seed for sampling.",
    )
    parser.add_argument(
        "--quantlib-source-root",
        type=Path,
        default=ROOT / "third_party_source" / "quantlib_copy" / "QuantLib",
        help="QuantLib source root containing ql/ headers.",
    )
    parser.add_argument(
        "--quantlib-build-dir",
        type=Path,
        default=ROOT / "third_party_source" / "quantlib_copy" / "build",
        help="QuantLib build root.",
    )
    parser.add_argument(
        "--quantlib-wrapper-src",
        type=Path,
        default=ROOT / "python" / "examples" / "quantlib_bsm_ref.cpp",
        help="C++ source for QuantLib wrapper.",
    )
    parser.add_argument(
        "--quantlib-wrapper-lib",
        type=Path,
        default=ROOT / "third_party_source" / "quantlib_copy" / "build" / "libql_ref_bsm.so",
        help="Output path for compiled QuantLib wrapper library.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    start = time.perf_counter()

    call_batch, put_batch, ingest_stats = _collect_batches(
        xlsx_path=args.xlsx,
        sheet_name=args.sheet,
        dte_basis=args.dte_basis,
        carry_mode=args.carry_mode,
        risk_free_rate=args.risk_free_rate,
        dividend_yield=args.dividend_yield,
        pcp_q_abs_max=args.pcp_q_abs_max,
        sample_per_side=args.sample_per_side,
        seed=args.seed,
    )

    _compile_quantlib_wrapper(
        wrapper_src=args.quantlib_wrapper_src,
        wrapper_lib=args.quantlib_wrapper_lib,
        quantlib_source_root=args.quantlib_source_root,
        quantlib_build_dir=args.quantlib_build_dir,
    )
    ql_ref = QuantLibBSMRef(args.quantlib_wrapper_lib)
    qk = QuantKernel()

    call_report = _run_side(
        qk=qk,
        ql_ref=ql_ref,
        side_name="call",
        batch=call_batch,
    )
    put_report = _run_side(
        qk=qk,
        ql_ref=ql_ref,
        side_name="put",
        batch=put_batch,
    )

    elapsed = time.perf_counter() - start

    report = {
        "config": {
            "xlsx": str(args.xlsx),
            "sheet": args.sheet if args.sheet else "first_sheet",
            "dte_basis": float(args.dte_basis),
            "carry_mode": args.carry_mode,
            "risk_free_rate": float(args.risk_free_rate),
            "dividend_yield": float(args.dividend_yield),
            "pcp_q_abs_max": float(args.pcp_q_abs_max),
            "sample_per_side": int(args.sample_per_side),
            "seed": int(args.seed),
            "quantlib_wrapper_lib": str(args.quantlib_wrapper_lib),
        },
        "ingest": ingest_stats,
        "call": call_report,
        "put": put_report,
        "total_runtime_seconds": float(elapsed),
    }
    print(json.dumps(report, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
